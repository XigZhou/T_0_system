from __future__ import annotations

import unittest

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from overnight_bt.backtest import LoadedSymbol, export_backtest_table_excel, run_portfolio_backtest_loaded
from overnight_bt.models import BacktestRequest, SignalQualityRequest
from overnight_bt.signal_quality import run_signal_quality_loaded
from tests.helpers import make_processed_stock


def _loaded_symbol(frame: pd.DataFrame) -> LoadedSymbol:
    clean = frame.reset_index(drop=True)
    return LoadedSymbol(
        symbol=str(clean.iloc[0]["symbol"]).strip().zfill(6),
        name=str(clean.iloc[0]["name"]).strip(),
        df=clean,
        idx_by_date={date: idx for idx, date in enumerate(clean["trade_date"].astype(str).tolist())},
    )


def _adj_factor_changed_frame() -> pd.DataFrame:
    return make_processed_stock(
        "000001",
        "TestCo",
        [
            {
                "trade_date": "20240102",
                "raw_open": 9.8,
                "raw_high": 10.2,
                "raw_low": 9.7,
                "raw_close": 10.0,
                "adj_factor": 1.0,
                "m20": 0.2,
                "can_buy_t": True,
                "can_buy_open_t": True,
                "can_sell_t": True,
                "can_sell_t1": True,
            },
            {
                "trade_date": "20240103",
                "raw_open": 10.0,
                "raw_high": 10.4,
                "raw_low": 9.9,
                "raw_close": 10.0,
                "adj_factor": 1.0,
                "m20": 0.2,
                "can_buy_t": True,
                "can_buy_open_t": True,
                "can_sell_t": True,
                "can_sell_t1": True,
            },
            {
                "trade_date": "20240104",
                "raw_open": 11.0,
                "raw_high": 12.2,
                "raw_low": 10.8,
                "raw_close": 12.0,
                "adj_factor": 1.2,
                "m20": 0.2,
                "can_buy_t": True,
                "can_buy_open_t": True,
                "can_sell_t": True,
                "can_sell_t1": True,
            },
        ],
    )


class RealShareSettlementTest(unittest.TestCase):
    def test_account_sell_uses_real_integer_shares_when_adj_factor_changes(self) -> None:
        loaded = [_loaded_symbol(_adj_factor_changed_frame())]

        result = run_portfolio_backtest_loaded(
            loaded,
            {"data_source": "unit-test"},
            BacktestRequest(
                stock_pool_template_name="unit-test",
                start_date="20240102",
                end_date="20240102",
                buy_condition="m20>0",
                score_expression="m20",
                top_n=1,
                initial_cash=100000.0,
                per_trade_budget=1000.0,
                lot_size=100,
                buy_fee_rate=0.0,
                sell_fee_rate=0.0,
                stamp_tax_sell=0.0,
                settlement_mode="complete",
                exit_mode="fixed",
                entry_offset=1,
                exit_offset=2,
                realistic_execution=True,
                slippage_bps=0.0,
                min_commission=0.0,
            ),
        )

        sell_trade = next(row for row in result["trade_rows"] if row["action"] == "SELL")

        self.assertEqual(sell_trade["shares"], 100)
        self.assertEqual(sell_trade["gross_amount"], 1100.0)
        self.assertEqual(sell_trade["pnl"], 100.0)

    def test_account_mark_to_market_uses_real_shares_when_adj_factor_changes(self) -> None:
        loaded = [_loaded_symbol(_adj_factor_changed_frame())]

        result = run_portfolio_backtest_loaded(
            loaded,
            {"data_source": "unit-test"},
            BacktestRequest(
                stock_pool_template_name="unit-test",
                start_date="20240102",
                end_date="20240104",
                buy_condition="m20>0",
                score_expression="m20",
                top_n=1,
                initial_cash=100000.0,
                per_trade_budget=1000.0,
                lot_size=100,
                buy_fee_rate=0.0,
                sell_fee_rate=0.0,
                stamp_tax_sell=0.0,
                settlement_mode="cutoff",
                exit_mode="fixed",
                entry_offset=1,
                exit_offset=3,
                realistic_execution=True,
                slippage_bps=0.0,
                min_commission=0.0,
            ),
        )

        last_daily_row = result["daily_rows"][-1]
        open_position = result["open_position_rows"][0]

        self.assertEqual(last_daily_row["market_value"], 1200.0)
        self.assertEqual(result["summary"]["ending_market_value"], 1200.0)
        self.assertEqual(open_position["shares"], 100)
        self.assertEqual(open_position["market_value"], 1200.0)

    def test_signal_quality_sell_uses_fixed_100_real_shares_when_adj_factor_changes(self) -> None:
        loaded = [_loaded_symbol(_adj_factor_changed_frame())]

        result = run_signal_quality_loaded(
            loaded,
            {"data_source": "unit-test"},
            SignalQualityRequest(
                stock_pool_template_name="unit-test",
                start_date="20240102",
                end_date="20240102",
                buy_condition="m20>0",
                score_expression="m20",
                top_n=1,
                buy_fee_rate=0.0,
                sell_fee_rate=0.0,
                stamp_tax_sell=0.0,
                settlement_mode="complete",
                exit_mode="fixed",
                entry_offset=1,
                exit_offset=2,
                realistic_execution=True,
                slippage_bps=0.0,
                min_commission=0.0,
            ),
        )

        sell_trade = next(row for row in result["trade_rows"] if row["action"] == "SELL")

        self.assertEqual(sell_trade["shares"], 100)
        self.assertEqual(sell_trade["gross_amount"], 1100.0)
        self.assertEqual(sell_trade["pnl"], 100.0)


    def test_account_result_declares_auditable_real_account_basis(self) -> None:
        loaded = [_loaded_symbol(_adj_factor_changed_frame())]

        result = run_portfolio_backtest_loaded(
            loaded,
            {"data_source": "unit-test"},
            BacktestRequest(
                stock_pool_template_name="unit-test",
                start_date="20240102",
                end_date="20240104",
                buy_condition="m20>0",
                score_expression="m20",
                top_n=1,
                initial_cash=100000.0,
                per_trade_budget=1000.0,
                lot_size=100,
                buy_fee_rate=0.0,
                sell_fee_rate=0.0,
                stamp_tax_sell=0.0,
                settlement_mode="cutoff",
                exit_mode="fixed",
                entry_offset=1,
                exit_offset=3,
                realistic_execution=True,
                slippage_bps=0.0,
                min_commission=0.0,
            ),
        )

        summary = result["summary"]
        self.assertEqual(summary["result_mode"], "account")
        self.assertEqual(summary["metric_basis"], "real_account")
        self.assertEqual(summary["equity_label"], "账户权益")
        self.assertEqual(summary["trade_flow_basis"], "真实账户成交流水")
        self.assertTrue(summary["auditable"])

        buy_trade = next(row for row in result["trade_rows"] if row["action"] == "BUY")
        self.assertEqual(buy_trade["shares"], 100)
        self.assertEqual(buy_trade["gross_amount"], 1000.0)
        self.assertEqual(buy_trade["net_amount"], 1000.0)

    def test_signal_quality_declares_sample_flow_and_signal_nav_basis(self) -> None:
        loaded = [_loaded_symbol(_adj_factor_changed_frame())]

        result = run_signal_quality_loaded(
            loaded,
            {"data_source": "unit-test"},
            SignalQualityRequest(
                stock_pool_template_name="unit-test",
                start_date="20240102",
                end_date="20240102",
                buy_condition="m20>0",
                score_expression="m20",
                top_n=1,
                buy_fee_rate=0.0,
                sell_fee_rate=0.0,
                stamp_tax_sell=0.0,
                settlement_mode="complete",
                exit_mode="fixed",
                entry_offset=1,
                exit_offset=2,
                realistic_execution=True,
                slippage_bps=0.0,
                min_commission=0.0,
            ),
        )

        summary = result["summary"]
        self.assertEqual(summary["result_mode"], "signal_quality")
        self.assertEqual(summary["metric_basis"], "signal_return_diagnostic")
        self.assertEqual(summary["equity_label"], "信号净值")
        self.assertEqual(summary["trade_flow_basis"], "固定100股样本流水，不代表真实账户成交")
        self.assertFalse(summary["auditable"])
        shares = {row["shares"] for row in result["trade_rows"] if row["action"] in {"BUY", "SELL"}}
        self.assertEqual(shares, {100.0})

    def test_trade_flow_export_contains_basis_sheet_and_audit_columns(self) -> None:
        loaded = [_loaded_symbol(_adj_factor_changed_frame())]
        result = run_portfolio_backtest_loaded(
            loaded,
            {"data_source": "unit-test"},
            BacktestRequest(
                stock_pool_template_name="unit-test",
                start_date="20240102",
                end_date="20240102",
                buy_condition="m20>0",
                score_expression="m20",
                top_n=1,
                initial_cash=100000.0,
                per_trade_budget=1000.0,
                lot_size=100,
                buy_fee_rate=0.0,
                sell_fee_rate=0.0,
                stamp_tax_sell=0.0,
                settlement_mode="complete",
                exit_mode="fixed",
                entry_offset=1,
                exit_offset=2,
                realistic_execution=True,
                slippage_bps=0.0,
                min_commission=0.0,
            ),
        )

        workbook = load_workbook(BytesIO(export_backtest_table_excel(result, "trade_rows")))

        self.assertIn("口径说明", workbook.sheetnames)
        self.assertIn("真实交易流水", workbook.sheetnames)
        basis = {row[0].value: row[1].value for row in workbook["口径说明"].iter_rows(min_row=2, max_col=2)}
        self.assertEqual(basis["本表是否可用于账户审计"], "是")
        self.assertEqual(basis["统计口径"], "real_account")
        headers = [cell.value for cell in next(workbook["真实交易流水"].iter_rows(min_row=1, max_row=1))]
        for label in ["成交金额", "费用", "净金额", "交易后现金", "交易收益", "退出原因"]:
            self.assertIn(label, headers)
